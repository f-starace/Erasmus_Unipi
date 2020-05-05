import ezsheets                                                         # necessario per connettersi alla API di Google

import os                                                               # sys e os permettono di cambiare la working directory in quella contenente lo script (si evitano così errori relativi a un path sbagliato)       
import sys

from tqdm import tqdm                                                   # per avere una interfaccia grafica che mostra il progresso del loop

import re                                                               # necessario per effettuare alcune operazioni di riconscimento di Pattern di testo (Regular Expression)

from GatherID import GatherId                                           # Imports the gatherId function 
from colors import Color                                                # Imports the Color class (per i colori dell'interfaccia)
from ErasmusManagerScraper import EM_Scraper                            # Imports the Scraper

import openpyxl                                                         # Non ho trovato un modo semplice per comunicare con l'API di Google per cambiare il colore delle celle, pertanto sposto i risultati su un foglio di calcolo excel
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import GREEN, RED





print(Color.bold + Color.underline +'\nSI CONSIGLIA DI UTILIZZARE QUESTO SCRIPT SU UNA COPIA DEL FOGLIO GOOGLE ORIGINALE' + Color.end)
print('\nPer bug o risoluzione dei problemi scrivere mail a: staracefederico@icloud.com')
input('\nPremi invio per continuare:      ')


#Changing the Current Working Directory to the one containing the script
os.chdir(sys.path[0])



# Retrive Spreadsheet ID from URL

url = input('Inserisci l\'url contenente la tabella [o premi solo invio per default]\n')
if url == '':
    url = 'https://docs.google.com/spreadsheets/d/1DoMqVZql9YTRED2Cb4uZkSm_ILZvrewzxELShvfTaC8/edit#gid=0'
    print(f'Utilizzo l\'url di test:    {url}\n')
    
try:
    SpreadSheetID = GatherId(url)
except:
    print(Color.red +'\nErrore, non sono riuscito a ricavare l\'ID dal link. Assicurati che sia corretto\n'+ Color.end)
    quit()




#opening the spreadsheet

print('Apro la tabella...')
ss = ezsheets.Spreadsheet(SpreadSheetID)
print(f'"{ss.title}" aperta\n')

choice = input('Vuoi avviare il processo di scraping delle mete erasmus? [y/n] (Invio = n):     ')

if choice == 'y':
    EM_Scraper(ss)

    
    





# Opening first sheet (which contains the ranking)
s = ss[0]


# Opening last sheet (which contains all the info gathered from ErasmusManager)
s2 = ss[-1]





def SeatsAviable(sheet, code_column, seats_column):
    '''
    This function will create a dictionary for each destionation in which:
    Keys will be represented by the univerity code
    Values are the number of seats 
    e.g. {'E VALENCI01': 4 , 'A WIEN64': 2, etc...}
    
    '''
    d = {} #dictionary
    for index in range(2,100):                                 #Not including the first row
        if sheet[f'{code_column}{index}'] != '':               #Ignoring empty cells
            
            num_seats = int(sheet[f'{seats_column}{index}'])   #Converting seats to int data type (previously string)
            
            d[sheet[f'{code_column}{index}']] = num_seats      #To assign a value to a key => d[key] = value
    
    d['void'] = -1
            
   
    return d


d = SeatsAviable(s2,'B','D')
print('\nCreato dizionario per l\'assegnazione dei posti alle mete \n')





def IsMatricola(string):
    '''
    Boolean Function, it returns True if the string contains 6 digits. Else it will return False
    '''
    pattern = r'\d\d\d\d\d\d' 
    
    if re.match(pattern, string) != None:
        return True
    else:
        return False
    


def CodeCheck(dictionary, graduatory_sheet):
    '''
    Questa funzione controlla se c'è qualche discrepanza tra i codici università contenuti nel foglio delle graduatorie e il foglio contenente i codici di ErasmusManager
    '''
    
    print(Color.underline + '\nVerifica della corrispondenza dei codici del foglio graduatorie con quello del foglio ErasmusManager:\n'+ Color.end)
    input('Per poter funzionare lo script ha bisogno di una corrispondenza esatta tra i due codici.\nQuesto check aiuta nell\'individuazione di errori di battitura e nell\'identificazione di errori qualora il candidato avesse scelto una meta non rientrante tra quelle disponibili\nPremi invio per continuare: ')
    d_keys = list(d.keys())
    
    pattern = r'\d\d\d\d\d\d' 
    
    
    for i in range(1,116):
        
        matricola = graduatory_sheet[f'A{i}']
        
        if re.match(pattern, matricola) != None:
            
            first_choice = graduatory_sheet[f'M{i}']
            second_choice = graduatory_sheet[f'N{i}']
            third_choice = graduatory_sheet[f'O{i}']

            if first_choice not in d_keys:
                
                print(f'Non riesco ad identificare il codice nella riga {i}\nVerifica l\'errore')
                while True:
                    correction = input(f'{Color.red}{first_choice}{Color.end} => ')
                    
                    if correction in d_keys:
                        print(Color.green + 'L\'errore è stato risolto' + Color.end)
                        graduatory_sheet[f'M{i}'] = correction
                        break
                    
                    elif correction == '':
                        definitive_choice = input(f'Sei sicuro di voler eliminare {first_choice} [y/n]')
                        if definitive_choice == 'y':
                            print(f'{first_choice} eliminata dalla cella')
                            graduatory_sheet[f'M{i}'] = correction = ''    
                            break                
                    else:
                        print('Errore persiste, riprova')

            if second_choice not in d_keys and second_choice !="":
                
                print(f'Non riesco ad identificare il codice nella riga {i}\nVerifica l\'errore')
                while True:
                    correction = input(f'{Color.red}{second_choice}{Color.end} => ')
                    
                    if correction in d_keys:
                        print(Color.green + 'L\'errore è stato risolto' + Color.end)
                        graduatory_sheet[f'N{i}'] = correction                 
                        break
                    
                    elif correction == '':
                        definitive_choice = input(f'Sei sicuro di voler eliminare {second_choice} [y/n]')
                        if definitive_choice == 'y':
                            print(f'{second_choice} eliminata dalla cella')
                            graduatory_sheet[f'N{i}'] = ''
                            break
                            
                    else:
                        print('Errore persiste, riprova')
                        
                        
            
            
            if third_choice not in d_keys and third_choice !="":
                
                print(f'Non riesco ad identificare il codice nella riga {i}\nVerifica l\'errore')
                while True:                    
                    correction = input(f'{Color.red}{third_choice}{Color.end} => ')
                    
                    if correction in d_keys:
                        print(Color.green + 'L\'errore è stato risolto' + Color.end)
                        graduatory_sheet[f'O{i}'] = correction
                        break
                    
                    elif correction == '':
                        definitive_choice = input(f'Sei sicuro di voler eliminare {third_choice} [y/n]')
                        if definitive_choice =='y':
                            print(f'{third_choice} eliminata dalla cella')
                            graduatory_sheet[f'O{i}'] = ''
                            break          
                    else:
                        print('Errore persiste, riprova')
    
    print(Color.green+ '\nNessun Conflitto rilevato\n'+ Color.end)
    
    
      
    
    
CodeCheck(d,s)




#OPENPYXL SETUP

input('Verrà creato un file in formato .xlsx all\'interno della cartella dello script dove saranno condivisi i risultati ottenuti\nPremi invio per continuare:    ')

#Importo i valori necessari per cambiare automaticamente il colore delle celle
greenFill = PatternFill(fgColor=GREEN, fill_type = "solid")
redFill = PatternFill(fgColor=RED, fill_type = "solid")





wb = openpyxl.Workbook()
# Creo un foglio dove verranno conservate le graduatorie:
g_s = wb['Sheet']
g_s.title = 'Graduatorie'




  
                    




print('Looping through each row and assigning seats to candidates')


for i in tqdm(range(2,116)):   
    # Looping through each row
    '''
    Questa parte è leggermente confusionaria, in pratica:
    1) Prendo il valore corrispondente all'indice da Google Sheets
    2) Assegno allo stesso indice la cella che mi interessa per trascrivere i dati su excel
    3) Assegno il valore preso da Google Sheets alla cella Excel
    
    Probabilmente si poteva fare in modo più elegante, ma non mi è venuto in mente niente di più immediato
    '''
    
    matricola = s[f'A{i}']
    matricola_excel = g_s[f'A{i}']
    matricola_excel.value = matricola
    
    
    prima_meta = s[f'M{i}']
    prima_meta_excel = g_s[f'B{i}']
    prima_meta_excel.value = prima_meta
     
    
    seconda_meta = s[f'N{i}']
    seconda_meta_excel =  g_s[f'C{i}']
    seconda_meta_excel.value = seconda_meta

    
    terza_meta = s[f'O{i}']
    terza_meta_excel =  g_s[f'D{i}']
    terza_meta_excel.value = terza_meta

    summary_excel= g_s[f'F{i}']
    


    
    if IsMatricola(matricola) == False:       # Se la cella non contiene una matricola viene ignorata dal loop
        continue
    
    
        
    elif d[prima_meta] > 0:                   # Controllo se ci sono posti disponibili per la meta grazie al dizionario
        prima_meta_excel.fill = greenFill     # Se si, la cella verrà colorata di verde 
                        
        d[prima_meta] -= 1                    # Rimuovo un posto dalla meta
        
        summary_excel.value = f'Entrato nella prima scelta a {prima_meta}, Posti rimasti {d[prima_meta]}'
    

        

    elif d[seconda_meta] > 0:   
    
        prima_meta_excel.fill = redFill
        seconda_meta_excel.fill = greenFill
        
        d[seconda_meta] -= 1                 # Rimuovo un posto dalla meta
        
        summary_excel.value= f'Entrato nella seconda scelta a {seconda_meta}, Posti rimasti {d[seconda_meta]}'

    
    elif d.get(seconda_meta) == -1:          # il valore -1 è stato assegnato alle key che ha come valore 'void' (cella vuota)
        
        prima_meta_excel.fill = redFill
        seconda_meta_excel.fill = redFill
        
        summary_excel.value= 'Tutte le sedi sono state prese'
        
    elif d[terza_meta] > 0:   
        
        prima_meta_excel.fill = redFill
        seconda_meta_excel.fill = redFill
        terza_meta_excel.fill = greenFill 
        
        d[terza_meta] -= 1                   # Rimuovo un posto dalla meta
        
        summary_excel.value= f'Entrato nella terza scelta a {terza_meta}, Posti rimasti {d[terza_meta]}'       
    
    else:
        prima_meta_excel.fill = redFill
        seconda_meta_excel.fill = redFill
        terza_meta_excel.fill = redFill 
        
        
        summary_excel.value= 'Tutte le sedi sono state prese'
        

# Creo un foglio dove verranno conservati i posti rimanenti della graduatoria:
m_s = wb.create_sheet("Posti rimanenti")

m_s['A1'] = 'SEDE'
m_s['B1'] = 'POSTI RIMASTI'

i = 2
for key, value in d.items():
    m_s[f'A{i}'] = key
    m_s[f'B{i}'] = value
    i += 1               
wb.save('results.xlsx')           

input('\n\n' + Color.green + 'Completato! Premi invio per terminare il programma   ' + Color.end)


